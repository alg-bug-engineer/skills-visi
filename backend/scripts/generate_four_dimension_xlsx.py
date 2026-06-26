#!/usr/bin/env python3
"""Generate four-dimension intersection filter Excel (morning/evening/off-peak + SQL)."""

from __future__ import annotations

import asyncio
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from intersection_agent.config import get_settings
from intersection_agent.db.postgres import PostgresPool
from intersection_agent.models.domain import TimePeriod
from intersection_agent.utils.data_window import build_data_window

OUTPUT = ROOT.parent / "docs/sql_queries/路口四维筛选结果.xlsx"
SQL_OUT = ROOT / "scripts/filter_four_dimension_intersections.sql"

PERIODS = [
    ("早高峰", TimePeriod(start="07:00", end="09:00", label="早高峰")),
    ("平峰", TimePeriod(start="10:00", end="16:00", label="平峰")),
    ("晚高峰", TimePeriod(start="17:00", end="19:00", label="晚高峰")),
]

HEADERS_ZH = [
    "路口ID",
    "路口名称",
    "分析时段",
    "时钟区间",
    "step_index区间",
    "DWS星期模式",
    "DWD日期起",
    "DWD日期止",
    "流量统计日起",
    "流量统计日止",
    "饱和度最大值",
    "失衡系数",
    "服务水平",
    "绿灯利用率最小值",
    "绿灯利用率均值",
    "最大排队(m)",
    "平均停车(s)",
    "延误指数均值",
    "平均停车次数",
    "转向流量合计",
    "转向流量均值",
    "进口通过率均值",
    "周期(s)",
    "方案数",
    "绿信比",
    "日计划时段数",
    "所属干线",
    "协调走廊",
]

ROW_KEYS = [
    "inter_id",
    "inter_name",
    "period_label",
    "time_slot",
    "step_range",
    "dws_dow",
    "dwd_date_from",
    "dwd_date_to",
    "flow_date_from",
    "flow_date_to",
    "sat_max",
    "imb_max",
    "los",
    "gu_min",
    "gu_avg",
    "q_max_m",
    "avg_stop_s",
    "avg_delay_idx",
    "stop_times_avg",
    "turn_flow_sum",
    "turn_flow_avg",
    "pass_flow_avg",
    "cycle_s",
    "plan_cnt",
    "green_ratio",
    "period_cnt",
    "line_names",
    "in_corridor",
]

STRING_COLS = {
    "inter_id",
    "inter_name",
    "period_label",
    "time_slot",
    "step_range",
    "dws_dow",
    "los",
    "line_names",
    "in_corridor",
}
INT_COLS = {"plan_cnt", "period_cnt", "turn_flow_sum"}
DATE_COLS = {"dwd_date_from", "dwd_date_to", "flow_date_from", "flow_date_to"}

COLUMN_WIDTHS = [
    18, 34, 10, 12, 14, 16, 12, 12, 12, 12,
    12, 10, 8, 14, 14, 12, 12, 12, 12,
    12, 12, 12, 8, 8, 8, 12, 36, 10,
]


def period_sql(
    period_label: str,
    step_start: int,
    step_end: int,
    slot_start: str,
    slot_end: str,
) -> str:
    step_range = f"{step_start}-{step_end}"
    time_slot = f"{slot_start}-{slot_end}"
    return f"""
WITH dwd_calendar AS (
    SELECT MIN(stat_time::date) AS dwd_date_from,
           MAX(stat_time::date) AS dwd_date_to
    FROM {{fs}}.dwd_tfc_inter_dir_perf_5min
    WHERE is_deleted = 0
      AND EXTRACT(DOW FROM stat_time) = 5
      AND stat_time::time >= '{slot_start}'::time
      AND stat_time::time < '{slot_end}'::time
),
eval_metrics AS (
    SELECT e.inter_id,
        ROUND(MAX(e.saturation_max)::numeric, 3) AS sat_max,
        ROUND(MAX(e.unbalance_index)::numeric, 3) AS imb_max,
        MAX(e.level_of_service) AS los
    FROM {{fs}}.dws_inter_evaluation_5min_mm e
    WHERE e.is_deleted = 0
      AND e.day_of_week = 5
      AND e.step_index BETWEEN {step_start} AND {step_end}
    GROUP BY e.inter_id
),
green_metrics AS (
    SELECT gu.inter_id,
        ROUND(AVG(gu.green_utilization)::numeric, 3) AS gu_avg,
        ROUND(MIN(gu.green_utilization)::numeric, 3) AS gu_min
    FROM {{fs}}.dws_turn_green_utilization_5min_mm gu
    WHERE gu.is_deleted = 0
      AND gu.day_of_week = 5
      AND gu.step_index BETWEEN {step_start} AND {step_end}
    GROUP BY gu.inter_id
),
turn_spread AS (
    SELECT t.inter_id,
        ROUND((MAX(t.turn_saturation) - MIN(t.turn_saturation))::numeric, 3) AS turn_spread
    FROM {{fs}}.dws_turn_saturation_5min_mm t
    WHERE t.is_deleted = 0
      AND t.day_of_week = 5
      AND t.step_index BETWEEN {step_start} AND {step_end}
    GROUP BY t.inter_id
),
dwd_metrics AS (
    SELECT d.inter_id,
        ROUND(MAX(d.queue_len_max)::numeric, 1) AS q_max_m,
        ROUND(AVG(d.stop_time)::numeric, 1) AS avg_stop_s,
        ROUND(AVG(d.delay_index)::numeric, 2) AS avg_delay_idx
    FROM {{fs}}.dwd_tfc_inter_dir_perf_5min d
    WHERE d.is_deleted = 0
      AND EXTRACT(DOW FROM stat_time) = 5
      AND d.stat_time::time >= '{slot_start}'::time
      AND d.stat_time::time < '{slot_end}'::time
    GROUP BY d.inter_id
),
flow_metrics AS (
    SELECT f.inter_id,
        MIN(f.flow_date_start) AS flow_date_from_raw,
        MAX(f.flow_date_end) AS flow_date_to_raw,
        ROUND(SUM(f.turn_flow_total)::numeric, 0) AS turn_flow_sum,
        ROUND(AVG(f.turn_flow_total)::numeric, 1) AS turn_flow_avg
    FROM {{fs}}.dws_inter_link_turn_flow_5min_mm f
    WHERE f.is_deleted = 0
      AND f.day_of_week = 5
      AND f.step_index BETWEEN {step_start} AND {step_end}
    GROUP BY f.inter_id
),
turn_perf AS (
    SELECT p.inter_id,
        ROUND(AVG(p.pass_flow)::numeric, 1) AS pass_flow_avg,
        ROUND(AVG(p.stop_times)::numeric, 2) AS stop_times_avg
    FROM {{fs}}.dws_inter_dir_turn_perf_5min_mm p
    WHERE p.is_deleted = 0
      AND p.turn_dir_no = 0
      AND p.day_of_week = 5
      AND p.step_index BETWEEN {step_start} AND {step_end}
    GROUP BY p.inter_id
),
ctl_metrics AS (
    SELECT p.inter_id,
        ROUND(AVG(p.cycle_len_sec)::numeric, 0) AS cycle_s,
        COUNT(DISTINCT p.plan_no) AS plan_cnt,
        ROUND(AVG(t.green_sec::float / NULLIF(p.cycle_len_sec, 0))::numeric, 3) AS green_ratio
    FROM {{fs}}.dwd_ctl_inter_plan_cfg p
    JOIN {{fs}}.dwd_ctl_inter_plan_stage_timing t
      ON p.inter_id = t.inter_id
     AND p.plan_no = t.plan_no
     AND t.is_deleted = 0
    WHERE p.is_deleted = 0
    GROUP BY p.inter_id
),
period_metrics AS (
    SELECT inter_id, COUNT(DISTINCT period_seq_no) AS period_cnt
    FROM {{fs}}.dwd_ctl_inter_day_plan_period
    WHERE is_deleted = 0
    GROUP BY inter_id
),
line_metrics AS (
    SELECT r.inter_id,
        STRING_AGG(DISTINCT l.line_name, '；' ORDER BY l.line_name) AS line_names
    FROM {{rs}}.dim_line_inter_rltn r
    JOIN {{rs}}.dim_line_info l ON l.line_id = r.line_id
    WHERE r.is_deleted = 0
    GROUP BY r.inter_id
),
joined AS (
    SELECT i.inter_id,
        i.inter_name,
        '{period_label}'::text AS period_label,
        '{time_slot}'::text AS time_slot,
        '{step_range}'::text AS step_range,
        '周五(day_of_week=5)'::text AS dws_dow,
        dc.dwd_date_from,
        dc.dwd_date_to,
        CASE
            WHEN fm.flow_date_from_raw IS NOT NULL
            THEN TO_CHAR(TO_DATE(fm.flow_date_from_raw::text, 'YYYYMMDD'), 'YYYY-MM-DD')
        END AS flow_date_from,
        CASE
            WHEN fm.flow_date_to_raw IS NOT NULL
            THEN TO_CHAR(TO_DATE(fm.flow_date_to_raw::text, 'YYYYMMDD'), 'YYYY-MM-DD')
        END AS flow_date_to,
        em.sat_max,
        em.imb_max,
        em.los,
        gm.gu_min,
        gm.gu_avg,
        dm.q_max_m,
        dm.avg_stop_s,
        dm.avg_delay_idx,
        tp.stop_times_avg,
        fm.turn_flow_sum,
        fm.turn_flow_avg,
        tp.pass_flow_avg,
        cm.cycle_s,
        cm.plan_cnt,
        cm.green_ratio,
        pm.period_cnt,
        lm.line_names,
        CASE WHEN EXISTS (
            SELECT 1 FROM {{fs}}.dws_corridor_coord_group g
            WHERE g.is_deleted = 0
              AND g.inter_ids_json::text LIKE '%' || i.inter_id || '%'
        ) THEN '是' ELSE '否' END AS in_corridor,
        CASE WHEN em.sat_max >= 0.80 THEN 1 ELSE 0 END AS hit_sat,
        CASE WHEN em.imb_max >= 0.30 OR ts.turn_spread >= 0.60 THEN 1 ELSE 0 END AS hit_imb,
        CASE WHEN gm.gu_min < 0.60 OR gm.gu_avg < 0.60 THEN 1 ELSE 0 END AS hit_empty,
        CASE WHEN dm.q_max_m >= 100 THEN 1 ELSE 0 END AS hit_spill
    FROM {{rs}}.dim_inter_info i
    JOIN eval_metrics em ON em.inter_id = i.inter_id
    CROSS JOIN dwd_calendar dc
    LEFT JOIN green_metrics gm ON gm.inter_id = i.inter_id
    LEFT JOIN turn_spread ts ON ts.inter_id = i.inter_id
    LEFT JOIN dwd_metrics dm ON dm.inter_id = i.inter_id
    LEFT JOIN flow_metrics fm ON fm.inter_id = i.inter_id
    LEFT JOIN turn_perf tp ON tp.inter_id = i.inter_id
    LEFT JOIN ctl_metrics cm ON cm.inter_id = i.inter_id
    LEFT JOIN period_metrics pm ON pm.inter_id = i.inter_id
    LEFT JOIN line_metrics lm ON lm.inter_id = i.inter_id
    WHERE i.version_id = '20260501'
      AND i.is_signalized = 1
)
SELECT {", ".join(ROW_KEYS)}
FROM joined
WHERE (hit_sat + hit_imb + hit_empty + hit_spill) >= 1
ORDER BY sat_max DESC NULLS LAST
"""


def build_master_sql() -> str:
    lines = [
        "-- 路口四维筛选 SQL（早高峰 / 平峰 / 晚高峰）",
        "-- 治理落脚点：信控配时与流量不匹配",
        "-- 输出：运行指标 + 流量/延误/信控/干线，不含命中判定列（WHERE 仍按四维筛选）",
        "--",
        "-- DWS：day_of_week=5 周五周模式，无日历 dt；DWD：stat_time 日历明细",
        "-- DWD 当前库日历：2026-06-08 ~ 2026-06-14（周五切片主要为 2026-06-12）",
        "--",
        "SET search_path TO road6, xianchang, public;",
        "",
    ]
    for sheet_name, tp in PERIODS:
        window = build_data_window(tp, reference_date=None)
        lines.append(
            f"-- ========== {sheet_name} {tp.start}-{tp.end} "
            f"step {window.step_start}-{window.step_end} =========="
        )
        lines.append(
            period_sql(sheet_name, window.step_start, window.step_end, tp.start, tp.end).strip()
        )
        lines.append("")
    return "\n".join(lines)


def _format_cell(key: str, val: object) -> object:
    if val is None:
        return None
    if key in DATE_COLS:
        if isinstance(val, datetime):
            return val.date().isoformat()
        if isinstance(val, date):
            return val.isoformat()
        return str(val)
    if key in STRING_COLS:
        return str(val)
    if key in INT_COLS:
        return int(val)
    if isinstance(val, (int, float)):
        return float(val) if key not in INT_COLS else int(val)
    return val


def row_to_values(row: dict) -> list:
    return [_format_cell(key, row.get(key)) for key in ROW_KEYS]


def write_sheet(ws, rows: list[dict]) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    ws.append(HEADERS_ZH)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row_to_values(row))
    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


async def fetch_period(
    pool: PostgresPool,
    fs: str,
    rs: str,
    period_label: str,
    tp: TimePeriod,
) -> list[dict]:
    window = build_data_window(tp)
    sql = period_sql(period_label, window.step_start, window.step_end, tp.start, tp.end).format(
        fs=fs, rs=rs
    )
    return await pool.fetch(sql)


async def main() -> None:
    settings = get_settings()
    if settings.mock_db:
        print("MOCK_DB=1，跳过")
        return

    sql_text = build_master_sql()
    SQL_OUT.write_text(sql_text, encoding="utf-8")

    pool = PostgresPool()
    await pool.connect()
    fs, rs = settings.pg_flow_schema, settings.pgschema

    period_rows: dict[str, list[dict]] = {}
    for sheet_name, tp in PERIODS:
        rows = await fetch_period(pool, fs, rs, sheet_name, tp)
        period_rows[sheet_name] = rows
        window = build_data_window(tp)
        print(f"{sheet_name} {tp.start}-{tp.end} step {window.step_start}-{window.step_end}: {len(rows)} rows")

    await pool.close()

    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, _ in PERIODS:
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, period_rows[sheet_name])

    ws_sql = wb.create_sheet("查询SQL")
    ws_sql.column_dimensions["A"].width = 120
    mono = Font(name="Courier New", size=10)
    for i, line in enumerate(sql_text.splitlines(), start=1):
        cell = ws_sql.cell(row=i, column=1, value=line)
        cell.font = mono
        cell.alignment = Alignment(vertical="top")
    ws_sql.freeze_panes = "A1"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"saved: {OUTPUT}")


if __name__ == "__main__":
    asyncio.run(main())
