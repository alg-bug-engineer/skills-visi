#!/usr/bin/env python3
"""导出指定路口×进口×转向的全部流量溯源数据到 Excel（Sheet1 数据，Sheet2 SQL）。"""

from __future__ import annotations

import asyncio
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from intersection_agent.config import get_settings
from intersection_agent.db.postgres import PostgresPool
from intersection_agent.models.domain import TimePeriod
from intersection_agent.services.flow_trace_service import (
    day_labels_for_filter,
    period_type_from_label,
)
from intersection_agent.utils.data_window import build_data_window
from intersection_agent.utils.traffic_labels import DIR8_LABELS, turn_label

TARGET_INTER_ID = "011wwe28ctu00001"
TARGET_INTER_NAME = "奥体西路与经十路路口"
F_DIR8 = 6  # 西进口
TURN_DIR_NO = 2  # 直行
PERIOD_LABEL = "晚高峰"
OUTPUT = (
    ROOT.parent
    / "artifacts"
    / "upstream-correlate"
    / f"{TARGET_INTER_NAME}_西直行_流量溯源.xlsx"
)

DIR8_ENTRY = {
    0: "北进口",
    1: "东北进口",
    2: "东进口",
    3: "东南进口",
    4: "南进口",
    5: "西南进口",
    6: "西进口",
    7: "西北进口",
}

HEADERS = [
    "目标路口ID",
    "目标路口名称",
    "进口方位编码",
    "进口方位",
    "转向编码",
    "转向",
    "上游路口ID",
    "上游路口名称",
    "上游来流方位编码",
    "上游来流方位",
    "上游来流转向编码",
    "上游来流转向",
    "来流方向",
    "途经占比%",
    "上游经度",
    "上游纬度",
    "统计月份",
    "时段类型",
    "星期口径",
    "溯源类型",
]


def build_sql(flow_schema: str, road_schema: str, version_id: str) -> str:
  period_type = period_type_from_label(PERIOD_LABEL)
  window = build_data_window(TimePeriod(label=PERIOD_LABEL, start="17:00", end="19:00"))
  day_labels = day_labels_for_filter(window.dow_filter)
  day_list = ", ".join(f"'{d}'" for d in day_labels)
  return f"""-- 奥体西路与经十路路口 · 西进口直行 · 全部流量溯源上游路口
-- 目标路口: {TARGET_INTER_ID} ({TARGET_INTER_NAME})
-- 口径: {PERIOD_LABEL} / period_type={period_type} / day_of_week IN ({day_list})
-- version_id={version_id}

SELECT
    fc.inter_id              AS target_inter_id,
    tgt.inter_name           AS target_inter_name,
    fc.f_dir8_no             AS f_dir8_no,
    fc.turn_dir_no           AS turn_dir_no,
    fc.cor_inter_id          AS upstream_inter_id,
    cor.inter_name           AS upstream_inter_name,
    fc.cor_f_dir8_no         AS cor_f_dir8_no,
    fc.cor_turn_dir_no       AS cor_turn_dir_no,
    fc.flow_share_ratio      AS path_coverage_pct,
    ST_X(ST_GeomFromText(cor.geom_center)) AS upstream_lng,
    ST_Y(ST_GeomFromText(cor.geom_center)) AS upstream_lat,
    fc.month,
    fc.period_type,
    fc.day_of_week,
    fc.trace_type
FROM {flow_schema}.dws_tfc_inter_turn_flow_correlate_m fc
LEFT JOIN {road_schema}.dim_inter_info tgt
  ON tgt.inter_id = fc.inter_id AND tgt.version_id = '{version_id}'
LEFT JOIN {road_schema}.dim_inter_info cor
  ON cor.inter_id = fc.cor_inter_id AND cor.version_id = '{version_id}'
WHERE fc.inter_id = '{TARGET_INTER_ID}'
  AND fc.trace_type = 'UPSTREAM'
  AND fc.is_deleted = 0
  AND fc.period_type = '{period_type}'
  AND fc.day_of_week IN ({day_list})
  AND fc.f_dir8_no = {F_DIR8}
  AND fc.turn_dir_no = {TURN_DIR_NO}
  AND fc.month = (
      SELECT MAX(m.month)
      FROM {flow_schema}.dws_tfc_inter_turn_flow_correlate_m m
      WHERE m.inter_id = '{TARGET_INTER_ID}' AND m.is_deleted = 0
  )
ORDER BY fc.flow_share_ratio DESC, fc.cor_inter_id;
"""


async def fetch_rows() -> tuple[list[dict], str, str]:
    settings = get_settings()
    if settings.mock_db:
        raise RuntimeError("MOCK_DB=1，请关闭 mock 后连接真实库导出")

    flow_schema = settings.pg_flow_schema
    road_schema = settings.pgschema
    version_id = settings.pg_version_id
    period_type = period_type_from_label(PERIOD_LABEL)
    window = build_data_window(TimePeriod(label=PERIOD_LABEL, start="17:00", end="19:00"))
    day_labels = day_labels_for_filter(window.dow_filter)

    sql = f"""
        SELECT fc.inter_id, tgt.inter_name AS target_inter_name,
               fc.f_dir8_no, fc.turn_dir_no,
               fc.cor_inter_id, cor.inter_name AS cor_inter_name,
               fc.cor_f_dir8_no, fc.cor_turn_dir_no,
               fc.flow_share_ratio,
               ST_X(ST_GeomFromText(cor.geom_center)) AS cor_lng,
               ST_Y(ST_GeomFromText(cor.geom_center)) AS cor_lat,
               fc.month, fc.period_type, fc.day_of_week, fc.trace_type
        FROM {flow_schema}.dws_tfc_inter_turn_flow_correlate_m fc
        LEFT JOIN {road_schema}.dim_inter_info tgt
          ON tgt.inter_id = fc.inter_id AND tgt.version_id = $3
        LEFT JOIN {road_schema}.dim_inter_info cor
          ON cor.inter_id = fc.cor_inter_id AND cor.version_id = $3
        WHERE fc.inter_id = $1 AND fc.trace_type = 'UPSTREAM' AND fc.is_deleted = 0
          AND fc.period_type = $2
          AND fc.day_of_week = ANY($4::text[])
          AND fc.f_dir8_no = $5
          AND fc.turn_dir_no = $6
          AND fc.month = (
              SELECT MAX(m.month) FROM {flow_schema}.dws_tfc_inter_turn_flow_correlate_m m
              WHERE m.inter_id = $1 AND m.is_deleted = 0
          )
        ORDER BY fc.flow_share_ratio DESC, fc.cor_inter_id
    """
    pool = PostgresPool(settings)
    await pool.connect()
    try:
        records = await pool.fetch(
            sql,
            TARGET_INTER_ID,
            period_type,
            version_id,
            day_labels,
            F_DIR8,
            TURN_DIR_NO,
        )
        return [dict(r) for r in records], flow_schema, road_schema
    finally:
        await pool.close()


def row_to_values(r: dict) -> list:
    f_dir8 = int(r.get("f_dir8_no") or F_DIR8)
    turn = int(r.get("turn_dir_no") or TURN_DIR_NO)
    cor_d8 = int(r.get("cor_f_dir8_no") or 0)
    cor_turn = int(r.get("cor_turn_dir_no") or 2)
    cov = r.get("flow_share_ratio")
    return [
        r.get("inter_id") or TARGET_INTER_ID,
        r.get("target_inter_name") or TARGET_INTER_NAME,
        f_dir8,
        DIR8_ENTRY.get(f_dir8, DIR8_LABELS.get(f_dir8, "")),
        turn,
        {1: "左转", 2: "直行", 3: "右转", 4: "掉头"}.get(turn, str(turn)),
        r.get("cor_inter_id"),
        r.get("cor_inter_name"),
        cor_d8,
        DIR8_LABELS.get(cor_d8, str(cor_d8)),
        cor_turn,
        {1: "左转", 2: "直行", 3: "右转", 4: "掉头"}.get(cor_turn, str(cor_turn)),
        turn_label(cor_d8, cor_turn),
        round(float(cov), 2) if cov is not None else None,
        float(r["cor_lng"]) if r.get("cor_lng") is not None else None,
        float(r["cor_lat"]) if r.get("cor_lat") is not None else None,
        r.get("month"),
        r.get("period_type"),
        r.get("day_of_week"),
        r.get("trace_type"),
    ]


def write_xlsx(rows: list[dict], flow_schema: str, road_schema: str) -> Path:
    settings = get_settings()
    sql_text = build_sql(flow_schema, road_schema, settings.pg_version_id)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "溯源路口"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(HEADERS, 1):
        cell = ws_data.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, 2):
        for col, val in enumerate(row_to_values(r), 1):
            ws_data.cell(row=i, column=col, value=val)

    for col in range(1, len(HEADERS) + 1):
        ws_data.column_dimensions[get_column_letter(col)].width = 16
    ws_data.column_dimensions["B"].width = 28
    ws_data.column_dimensions["H"].width = 28

    # 汇总：distinct 上游路口数
    distinct = len({r.get("cor_inter_id") for r in rows if r.get("cor_inter_id")})
    ws_data.cell(row=len(rows) + 3, column=1, value=f"导出时间: {datetime.now():%Y-%m-%d %H:%M}")
    ws_data.cell(row=len(rows) + 4, column=1, value=f"明细行数: {len(rows)}")
    ws_data.cell(row=len(rows) + 5, column=1, value=f"distinct上游路口: {distinct}")

    ws_sql = wb.create_sheet("SQL")
    ws_sql.column_dimensions["A"].width = 120
    for i, line in enumerate(sql_text.strip().splitlines(), 1):
        ws_sql.cell(row=i, column=1, value=line)
        ws_sql.cell(row=i, column=1).alignment = Alignment(wrap_text=False, vertical="top")

    wb.save(OUTPUT)
    return OUTPUT


async def main() -> None:
    rows, flow_schema, road_schema = await fetch_rows()
    if not rows:
        print("未查到溯源数据，请检查路口/时段/库连接")
        sys.exit(1)
    path = write_xlsx(rows, flow_schema, road_schema)
    distinct = len({r.get("cor_inter_id") for r in rows if r.get("cor_inter_id")})
    print(f"已导出 {len(rows)} 行明细，{distinct} 个 distinct 上游路口")
    print(f"文件: {path}")


if __name__ == "__main__":
    asyncio.run(main())
